from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import io
import psycopg2
from psycopg2 import Error
import json
from openpyxl import Workbook
from io import BytesIO
import datetime
import re
from collections import defaultdict
from openpyxl.styles import Font, PatternFill

app = Flask(__name__)
app.config['SESSION_COOKIE_SECURE'] = True
DB_CONFIG = {
    "user": "tomer",
    "password": "t1",
    "host": "localhost",
    "port": "5433",
    "database": "appData"
}
def clean_value(value):
    if value == 'NaT' or value == 'nan' or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, datetime.datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    return value

def connect_to_db():
    try:
        connection = psycopg2.connect(**DB_CONFIG)
        connection.autocommit = True
        return connection
    except Error as e:
        print(f"Error while connecting to PostgreSQL: {e}")
        return None


def append_and_color_header(worksheet, headers, background_color):
    worksheet.append(headers)
    for cell in worksheet[worksheet.max_row]:
        cell.fill = PatternFill(start_color=background_color, end_color=background_color, fill_type="solid")
        cell.font = Font(color="000000")  # Black text
def process_flexible_data(headers, patient_code, data):
    # Create a defaultdict to store all values for each column
    data = sorted(data, key=lambda x: x[-2])
    data = [row[:-2] + row[-1:] for row in data]
    final_row = [patient_code] + list(data[0][:-1])

    # Append each last element of the rows to final_row
    for row in data:
        final_row.append(row[-1])

    return final_row



@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files and 'additionalDetails' not in request.form and 'selectedQuestions' not in request.form:
        return jsonify({'error': 'No file, data, or selected questions provided'})
    detail_type = request.form.get('detailType')


    file = request.files.get('file')
    data_text = request.form.get('additionalDetails')
    selected_questions = json.loads(request.form.get('selectedQuestions', '[]'))

    codes_array = []

    if file and file.filename:
        try:
            file_content = file.read()
            if file.filename.endswith('.csv'):
                df = pd.read_csv(io.StringIO(file_content.decode('utf-8')))
            elif file.filename.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(io.BytesIO(file_content))
            else:
                return jsonify({'error': 'Unsupported file format'})

            codes_array = df.values.flatten().tolist()
            # Regular expression pattern for matching paths or IDs
            pattern = re.compile(r'^([A-Za-z]:?(?:[\w-]+[/\\])*[\w-]+|[A-Za-z0-9_]+|[A-Za-z0-9]{6,})$')

            codes_array = [
                code for code in codes_array
                if (isinstance(code, str) and pattern.match(code)) or
                   (isinstance(code, (int, float)) and not pd.isna(code) and pattern.match(str(code)))
            ]
            if 'Uploaded' in codes_array:
                codes_array=codes_array[:-3]



        except Exception as e:
            return jsonify({'error': str(e)})

    elif data_text and len(codes_array)==0:
        codes_array.extend(data_text.split())

    elif not codes_array:
        return jsonify({'error': 'No valid patient codes provided'})

    elif not selected_questions:
        return jsonify({'error': 'No questions selected'})

    conn = connect_to_db()
    if not conn:
        return jsonify({'error': 'Unable to connect to the database'})

    try:
        with (conn.cursor() as cur):
            crf_columns = ['groupname', 'age', 'gender', 'scanno','datetimescan', 'heightcm', 'weightkg', 'study', 'condition','scanfile']
            question_ids = ', '.join(f"'{qid}'" for qid in selected_questions if qid not in crf_columns)

            if question_ids!="":
              cur.execute(f"SELECT questioneid, question FROM questiones WHERE questioneid IN ({question_ids})")
              categories = cur.fetchall()



            # Fetch answers
            results = {}
            answer_columns=['questioneid','answer']
            answer_questionsids = [str(i) for i in range(1, 501)]
            select_columns=[]
            left_joins = []
            flag=0
            headers = ["Patient Code"]
            for column in crf_columns:
                if column in selected_questions:
                    if column=='datetimescan' and detail_type=='pathScanFile':
                       select_columns.append(f"crf.{column}")
                    else:
                       select_columns.append(f"{column}")
                    headers=headers+[column]
            if detail_type=='subjectId' or detail_type=='pathScanFile':
            # Create Excel file
              wb = Workbook()
              ws = wb.active
              ws.title = "Results"
              # Write headers
              headers = ["Patient Code"] + [qid for qid in selected_questions if qid in crf_columns]
              conn = connect_to_db()
              with conn.cursor() as cur:
                if select_columns:
                  ws.append([""])
                  select_columns_str = ", ".join(select_columns)
                  headers=["Subject"]+["details"]+["at"]+["the"]+["time"]+["of"]+["the"]+["scan"]
                  append_and_color_header(ws, headers, "FFFFFF00")
                  ws.append([""])
                  ws.append([""])
                  if detail_type=='subjectId':
                    headers=["Subject ID"]+select_columns
                    ws.append(headers)
                  elif detail_type=='pathScanFile':
                    headers = ["Path Scan File"] + select_columns
                    ws.append(headers)
                  if detail_type=='subjectId':
                    for code in codes_array:
                      query = f""" SELECT {select_columns_str}
                                   FROM patients inner join crf on patients.patientcode=crf.patientcode
                                   WHERE patients.patientcode = ('{code}')
                               """
                      cur.execute(query)
                      result = cur.fetchone()
                      if result:
                          cleaned_result = [clean_value(value) for value in result]
                          ws.append([code]+cleaned_result)
                  elif detail_type=='pathScanFile':
                      for code in codes_array:
                          query = f""" SELECT {select_columns_str}
                                       FROM patients inner join crf on patients.patientcode=crf.patientcode inner join sharedscans on crf.datetimescan=sharedscans.datetimescan
                                       WHERE sharedscans.path = ('{code}')
                                 """
                          cur.execute(query)
                          result = cur.fetchone()
                          if result:
                              cleaned_result = [clean_value(value) for value in result]
                              ws.append([code] + cleaned_result)
                if question_ids!="":
                   cur.execute(f""" SELECT questioneid,question
                                    from questiones
                                    where questioneid IN ({question_ids}) order by questioneid""")
                   questions = cur.fetchall()
                   ws.append([""])
                   headers = ["Subject"] + ["details"] + ["from"] + ["questionaire"]
                   append_and_color_header(ws, headers, "FFFF0000")

                   result=""
                   if detail_type == 'subjectId':
                     headers = ["Subject ID"] + [question[1] for question in questions]
                     ws.append([""])
                     ws.append([""])
                     ws.append(headers)
                     for code in codes_array:
                       query = f""" SELECT answers.questioneid,answers.answer
                                    FROM patients inner join answers on patients.patientcode=answers.patientcode
                                    WHERE patients.patientcode = ('{code}') and answers.questioneid IN ({question_ids})
                                """
                       cur.execute(query)
                       result = cur.fetchall()
                       keys = [item[0] for item in result]
                       question_ids_temp = re.findall(r'\d+', question_ids)
                       # Convert the extracted strings to integers
                       question_ids_temp = [int(num) for num in question_ids_temp]
                       for question_id_temp in question_ids_temp:
                           if question_id_temp not in keys:
                               result.append((question_id_temp, 'Nan'))
                       processed_data = process_flexible_data(headers, code, result)
                       ws.append(processed_data)
                   elif detail_type == 'pathScanFile':
                     headers = ["Path Scan File"] + [question[1] for question in questions]
                     ws.append([""])
                     ws.append([""])
                     ws.append(headers)
                     for code in codes_array:
                        query = f""" SELECT answers.questioneid,answers.answer
                                        FROM patients inner join answers on patients.patientcode=answers.patientcode
                                        inner join crf on patients.patientcode=crf.patientcode inner join sharedscans on crf.datetimescan=sharedscans.datetimescan
                                        WHERE sharedscans.path = ('{code}') and answers.questioneid IN ({question_ids})
                                    """
                        cur.execute(query)
                        result = cur.fetchall()
                        keys = [item[0] for item in result]
                        question_ids_temp = re.findall(r'\d+', question_ids)
                        # Convert the extracted strings to integers
                        question_ids_temp = [int(num) for num in question_ids_temp]
                        for question_id_temp in question_ids_temp:
                          if question_id_temp not in keys:
                              result.append((question_id_temp, 'Nan'))
                        processed_data=process_flexible_data(headers, code, result)
                        ws.append(processed_data)
              excel_file = BytesIO()
              wb.save(excel_file)
              excel_file.seek(0)
              return send_file(
                      excel_file,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                      as_attachment=True,
                      download_name='results.xlsx'
              )
    except psycopg2.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred while processing questions'})
    finally:
        if conn:
            conn.close()


@app.route('/get_questions', methods=['GET'])
def get_questions():
    category = request.args.get('category')
    category_index_ranges = {
        'דמוגרפי כללי': [(3, 13)],
        'שפה ושיוך': [(14, 17)],
        'מצב משפחתי': [(17, 22)],
        'השכלה ומקצוע': [(23, 28)],
        'תחביבים והעדפות': [(29, 32), (464, 481), (489, 499)],
        'אורח חיים ועמדות': [(33, 47), (299, 312), (461, 464), (484, 484)],
        'שאלון שינה': [(48, 73)],
        'מצב בריאותי': [(74, 207), (290, 298)],
        'שאלון אישיות': [(208, 255)],
        'שאלון פסיכומטרי': [(256, 264)],
        'שאלון חרדה': [(265, 269)],
        'שאלון פוביות': [(270, 277)],
        'שאלון צאצאים שורדי שואה': [(278, 289)],
        'שאלון מוזיקה': [(313, 329)],
        'שאלון תכנות': [(330, 339)],
        'שאלון סמארטפון': [(340, 368)],
        'שאלון דיכאון וחרדה': [(369, 386)],
        'שאלון פוסט טראומה': [(387, 434)],
        'שאלון שבעה באוקטובר': [(435, 460)],
        'שאלות סיום': [(482, 483)],
        'All the questions': [(3, 501)],
    }

    conn = connect_to_db()
    if not conn:
        return jsonify({'error': 'Unable to connect to the database'})

    try:
        with conn.cursor() as cur:
            if category == 'Most_common_questions':
                common_questions = [
                    ('gender', 'Gender (at the time of the scan)'),
                    ('datetimescan', 'Date and time of scan'),
                    ('age', 'Age (at the time of the scan)'),
                    ('weightkg', 'Weight (kg) (at the time of the scan)'),
                    ('heightcm', 'Height (cm) (at the time of the scan)'),
                    ('study', 'study'),
                    ('condition', 'condition'),
                    ('groupname', 'Group'),
                    ('4', 'Dominant hand'),
                    #('Scan Details', 'Scan Details')
                ]
                cur.execute("SELECT * FROM questiones WHERE questioneid >= 14 AND questioneid <= 15")
                custom_questions = cur.fetchall()
                cur.execute("SELECT * FROM questiones WHERE questioneid >= 23 AND questioneid <= 28")
                education_work_questions = cur.fetchall()
                cur.execute("SELECT * FROM questiones WHERE questioneid >= 313 AND questioneid <= 329")
                music_questions = cur.fetchall()

                all_questions = common_questions + custom_questions + education_work_questions + music_questions
                return jsonify({'questions': all_questions})
            elif category == 'Patient_details_at_the_time_of_scan':
                patient_details = [
                    ('gender', 'Gender (at the time of the scan) '),
                    ('datetimescan', 'Date and time of scan'),
                    ('age', 'Age (at the time of the scan)'),
                    ('weightkg', 'Weight (kg) (at the time of the scan)'),
                    ('heightcm', 'Height (cm) (at the time of the scan)'),
                    ('study', 'study'),
                    ('condition', 'condition'),
                    ('groupname', 'Group'),
                    # ('Scan Details', 'Scan Details')
                ]
                return jsonify({'questions': patient_details})
            elif category in category_index_ranges:
                questions = []
                for start_index, end_index in category_index_ranges[category]:
                    cur.execute("SELECT * FROM questiones WHERE questioneid >= %s AND questioneid <= %s",
                                (start_index, end_index))
                    questions.extend(cur.fetchall())
                return jsonify({'questions': questions})
            else:
                return jsonify({'error': 'Invalid category'})
    except psycopg2.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'})
    finally:
        if conn:
            conn.close()


if __name__ == '__main__':
    app.run()